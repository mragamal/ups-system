import os
import html
import sqlite3
import hashlib
import shutil
from typing import Any

import openpyxl
from fastapi import FastAPI, Request, Form, File, UploadFile, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles

app = FastAPI(title="UPS System")

DB_FILE = "ups.db"
UPLOAD_DIR = "uploads"
LOGO_FILENAME = "logo.png"

os.makedirs(UPLOAD_DIR, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")

SECTIONS: dict[str, dict[str, Any]] = {
    "clients": {
        "title": "Clients",
        "table": "clients",
        "singular": "Client",
        "fields": [
            ("name", "Name", "text"),
            ("department", "Department", "text"),
            ("phone", "Phone", "text"),
            ("address", "Address", "text"),
        ],
    },
    "vendors": {
        "title": "Vendors",
        "table": "vendors",
        "singular": "Vendor",
        "fields": [
            ("name", "Name", "text"),
            ("phone", "Phone", "text"),
            ("address", "Address", "text"),
        ],
    },
    "accounts": {
        "title": "Accounts",
        "table": "accounts",
        "singular": "Account",
        "fields": [
            ("code", "Code", "text"),
            ("name", "Name", "text"),
            ("account_type", "Type", "text"),
            ("balance", "Balance", "number"),
        ],
    },
    "sites": {
        "title": "Sites",
        "table": "sites",
        "singular": "Site",
        "fields": [
            ("name", "Name", "text"),
            ("location", "Location", "text"),
            ("status", "Status", "text"),
        ],
    },
    "tickets": {
        "title": "Tickets",
        "table": "tickets",
        "singular": "Ticket",
        "fields": [
            ("title", "Title", "text"),
            ("client_name", "Client", "text"),
            ("site_name", "Site", "text"),
            ("status", "Status", "text"),
            ("description", "Description", "textarea"),
        ],
    },
    "work_orders": {
        "title": "Work Orders",
        "table": "work_orders",
        "singular": "Work Order",
        "fields": [
            ("ticket_id", "Ticket ID", "number"),
            ("description", "Description", "textarea"),
            ("status", "Status", "text"),
        ],
    },
    "trips": {
        "title": "Trips",
        "table": "trips",
        "singular": "Trip",
        "fields": [
            ("title", "Title", "text"),
            ("from_location", "From", "text"),
            ("to_location", "To", "text"),
            ("status", "Status", "text"),
            ("cost", "Cost", "number"),
        ],
    },
}


def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


def esc(value: Any) -> str:
    return html.escape("" if value is None else str(value))


def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode("utf-8")).hexdigest()


def init_db() -> None:
    conn = get_conn()
    cur = conn.cursor()

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            full_name TEXT DEFAULT '',
            password_hash TEXT NOT NULL,
            role TEXT DEFAULT 'user',
            is_active INTEGER DEFAULT 1
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            department TEXT DEFAULT '',
            phone TEXT DEFAULT '',
            address TEXT DEFAULT ''
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS vendors (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            phone TEXT DEFAULT '',
            address TEXT DEFAULT ''
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT NOT NULL,
            name TEXT NOT NULL,
            account_type TEXT DEFAULT '',
            balance REAL DEFAULT 0
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS sites (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            location TEXT DEFAULT '',
            status TEXT DEFAULT ''
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS tickets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            client_name TEXT DEFAULT '',
            site_name TEXT DEFAULT '',
            status TEXT DEFAULT '',
            description TEXT DEFAULT ''
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS work_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ticket_id INTEGER DEFAULT 0,
            description TEXT DEFAULT '',
            status TEXT DEFAULT ''
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS trips (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            from_location TEXT DEFAULT '',
            to_location TEXT DEFAULT '',
            status TEXT DEFAULT '',
            cost REAL DEFAULT 0
        )
        """
    )

    conn.commit()

    admin = conn.execute("SELECT id FROM users WHERE username = ?", ("admin",)).fetchone()
    if not admin:
        conn.execute(
            "INSERT INTO users (username, full_name, password_hash, role, is_active) VALUES (?, ?, ?, ?, ?)",
            ("admin", "System Admin", hash_password("admin123"), "admin", 1),
        )
        conn.commit()

    conn.close()


@app.on_event("startup")
def startup() -> None:
    init_db()


def current_user(request: Request):
    username = request.cookies.get("ups_user", "")
    if not username:
        return None
    conn = get_conn()
    row = conn.execute(
        "SELECT * FROM users WHERE username = ? AND is_active = 1",
        (username,),
    ).fetchone()
    conn.close()
    return row


def require_user(request: Request):
    user = current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Login required")
    return user


def require_admin(request: Request):
    user = require_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    return user


def section_exists(section: str) -> None:
    if section not in SECTIONS:
        raise HTTPException(status_code=404, detail="Section not found")


def count_rows(table_name: str) -> int:
    conn = get_conn()
    count = conn.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
    conn.close()
    return count


def logo_url() -> str:
    path = os.path.join(UPLOAD_DIR, LOGO_FILENAME)
    if os.path.exists(path):
        return f"/uploads/{LOGO_FILENAME}"
    return ""


def login_page_html(error: str = "") -> str:
    logo = logo_url()
    logo_html = f'<img src="{logo}" class="login-logo" alt="logo">' if logo else '<div class="login-logo-fallback">UPS</div>'
    error_html = f'<div class="login-error">{esc(error)}</div>' if error else ""

    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8"/>
        <meta name="viewport" content="width=device-width, initial-scale=1"/>
        <title>Login - UPS System</title>
        <style>
            * {{ box-sizing: border-box; }}
            body {{
                margin: 0;
                font-family: Arial, sans-serif;
                min-height: 100vh;
                background: linear-gradient(135deg, #0f172a, #1d4ed8);
                display: flex;
                align-items: center;
                justify-content: center;
                color: #111827;
            }}
            .login-shell {{
                width: 100%;
                max-width: 1000px;
                min-height: 560px;
                display: grid;
                grid-template-columns: 1.1fr .9fr;
                background: white;
                border-radius: 24px;
                overflow: hidden;
                box-shadow: 0 25px 60px rgba(0,0,0,.25);
            }}
            .login-visual {{
                background:
                    linear-gradient(rgba(15,23,42,.55), rgba(29,78,216,.65)),
                    url('{logo if logo else ""}') center/cover no-repeat,
                    linear-gradient(135deg, #1d4ed8, #0f172a);
                color: white;
                padding: 40px;
                display: flex;
                flex-direction: column;
                justify-content: space-between;
            }}
            .login-title {{
                font-size: 42px;
                font-weight: 700;
                line-height: 1.15;
            }}
            .login-sub {{
                font-size: 18px;
                opacity: .9;
                margin-top: 14px;
            }}
            .login-badge {{
                display: inline-block;
                background: rgba(255,255,255,.14);
                padding: 10px 14px;
                border-radius: 999px;
                font-size: 14px;
                margin-bottom: 18px;
            }}
            .login-card {{
                padding: 46px 38px;
                display: flex;
                flex-direction: column;
                justify-content: center;
            }}
            .login-logo {{
                max-width: 120px;
                max-height: 120px;
                object-fit: contain;
                margin-bottom: 18px;
            }}
            .login-logo-fallback {{
                width: 92px;
                height: 92px;
                border-radius: 20px;
                background: #e0e7ff;
                color: #1d4ed8;
                display: flex;
                align-items: center;
                justify-content: center;
                font-size: 30px;
                font-weight: 700;
                margin-bottom: 18px;
            }}
            .login-form-title {{
                font-size: 30px;
                font-weight: 700;
                margin-bottom: 8px;
            }}
            .login-form-sub {{
                color: #6b7280;
                margin-bottom: 24px;
            }}
            .field {{
                margin-bottom: 14px;
            }}
            .field label {{
                display: block;
                margin-bottom: 6px;
                font-weight: 700;
                color: #374151;
            }}
            .field input {{
                width: 100%;
                border: 1px solid #d1d5db;
                border-radius: 12px;
                padding: 13px 14px;
                font-size: 15px;
            }}
            .btn {{
                width: 100%;
                border: none;
                border-radius: 12px;
                padding: 14px;
                background: #2563eb;
                color: white;
                font-size: 16px;
                font-weight: 700;
                cursor: pointer;
                margin-top: 8px;
            }}
            .login-error {{
                background: #fee2e2;
                color: #991b1b;
                padding: 12px 14px;
                border-radius: 12px;
                margin-bottom: 16px;
            }}
            .hint {{
                margin-top: 18px;
                color: #6b7280;
                font-size: 14px;
            }}
            @media (max-width: 900px) {{
                .login-shell {{
                    grid-template-columns: 1fr;
                    margin: 20px;
                }}
                .login-visual {{
                    min-height: 220px;
                }}
            }}
        </style>
    </head>
    <body>
        <div class="login-shell">
            <div class="login-visual">
                <div>
                    <div class="login-badge">Ultra Power Solutions</div>
                    <div class="login-title">Welcome to UPS System</div>
                    <div class="login-sub">Enter your username and password to access the full management system.</div>
                </div>
                <div style="font-size:14px;opacity:.9;">Clients • Vendors • Accounts • Sites • Tickets • Work Orders • Trips</div>
            </div>
            <div class="login-card">
                {logo_html}
                <div class="login-form-title">Sign In</div>
                <div class="login-form-sub">Use your account to continue.</div>
                {error_html}
                <form method="post" action="/login">
                    <div class="field">
                        <label>Username</label>
                        <input type="text" name="username" required>
                    </div>
                    <div class="field">
                        <label>Password</label>
                        <input type="password" name="password" required>
                    </div>
                    <button class="btn" type="submit">Login</button>
                </form>
                <div class="hint">Default admin: admin / admin123</div>
            </div>
        </div>
    </body>
    </html>
    """


def nav_html(active: str = "dashboard", is_admin: bool = False) -> str:
    links = [
        ("dashboard", "Dashboard", "/ui"),
        ("clients", "Clients", "/ui/clients"),
        ("vendors", "Vendors", "/ui/vendors"),
        ("accounts", "Accounts", "/ui/accounts"),
        ("sites", "Sites", "/ui/sites"),
        ("tickets", "Tickets", "/ui/tickets"),
        ("work_orders", "Work Orders", "/ui/work_orders"),
        ("trips", "Trips", "/ui/trips"),
        ("logo", "Logo", "/ui/logo"),
    ]
    if is_admin:
        links.append(("users", "Users", "/ui/users"))

    items = []
    for key, label, href in links:
        cls = "nav-link active" if key == active else "nav-link"
        items.append(f'<a class="{cls}" href="{href}">{label}</a>')

    items.append('<a class="nav-link" href="/logout">Logout</a>')
    return "".join(items)


def page_html(title: str, body: str, user: sqlite3.Row, active: str = "dashboard") -> str:
    logo = logo_url()
    logo_html = f'<img src="{logo}" alt="logo">' if logo else '<div class="logo-fallback">UPS</div>'

    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8" />
        <meta name="viewport" content="width=device-width, initial-scale=1" />
        <title>{esc(title)}</title>
        <style>
            * {{ box-sizing: border-box; }}
            body {{
                margin: 0;
                font-family: Arial, sans-serif;
                background: #f3f6fb;
                color: #1f2937;
            }}
            .layout {{
                display: flex;
                min-height: 100vh;
            }}
            .sidebar {{
                width: 260px;
                background: #16335b;
                color: white;
                padding: 24px 18px;
            }}
            .logo-box {{
                background: white;
                border-radius: 18px;
                padding: 12px;
                text-align: center;
                margin-bottom: 18px;
            }}
            .logo-box img {{
                max-width: 100%;
                max-height: 90px;
                object-fit: contain;
            }}
            .logo-fallback {{
                height: 80px;
                display: flex;
                align-items: center;
                justify-content: center;
                font-size: 30px;
                font-weight: 700;
                color: #1d4ed8;
            }}
            .brand {{
                font-size: 26px;
                font-weight: 700;
                margin-bottom: 8px;
            }}
            .brand small {{
                display: block;
                font-size: 13px;
                font-weight: 400;
                opacity: .85;
                margin-top: 6px;
            }}
            .user-box {{
                background: rgba(255,255,255,.08);
                border-radius: 14px;
                padding: 12px;
                margin: 16px 0 18px 0;
                font-size: 14px;
            }}
            .nav-link {{
                display: block;
                color: white;
                text-decoration: none;
                padding: 12px 14px;
                border-radius: 10px;
                margin-bottom: 8px;
                background: rgba(255,255,255,.05);
            }}
            .nav-link:hover, .nav-link.active {{
                background: rgba(255,255,255,.15);
            }}
            .content {{
                flex: 1;
                padding: 28px;
            }}
            .page-title {{
                font-size: 32px;
                font-weight: 700;
                margin-bottom: 20px;
            }}
            .cards {{
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
                gap: 18px;
                margin-bottom: 24px;
            }}
            .card {{
                background: white;
                border-radius: 16px;
                padding: 22px;
                box-shadow: 0 8px 20px rgba(0,0,0,.06);
            }}
            .card-number {{
                font-size: 34px;
                font-weight: 700;
                margin-bottom: 8px;
            }}
            .toolbar {{
                display: flex;
                gap: 10px;
                margin-bottom: 16px;
                flex-wrap: wrap;
            }}
            .btn {{
                display: inline-block;
                border: none;
                border-radius: 10px;
                padding: 10px 14px;
                text-decoration: none;
                cursor: pointer;
                font-size: 14px;
            }}
            .btn-primary {{ background: #2563eb; color: white; }}
            .btn-warning {{ background: #f59e0b; color: white; }}
            .btn-danger {{ background: #dc2626; color: white; }}
            .btn-light {{ background: #e5e7eb; color: #111827; }}
            .btn-success {{ background: #16a34a; color: white; }}
            table {{
                width: 100%;
                border-collapse: collapse;
                background: white;
                border-radius: 16px;
                overflow: hidden;
                box-shadow: 0 8px 20px rgba(0,0,0,.06);
            }}
            th, td {{
                padding: 14px;
                text-align: left;
                border-bottom: 1px solid #e5e7eb;
                vertical-align: top;
            }}
            th {{
                background: #eef3fb;
            }}
            .form-box {{
                max-width: 820px;
                background: white;
                border-radius: 16px;
                padding: 24px;
                box-shadow: 0 8px 20px rgba(0,0,0,.06);
            }}
            .field {{
                margin-bottom: 14px;
            }}
            .field label {{
                display: block;
                margin-bottom: 6px;
                font-weight: 700;
            }}
            .field input, .field textarea, .field select {{
                width: 100%;
                border: 1px solid #cbd5e1;
                border-radius: 10px;
                padding: 10px 12px;
                font-size: 14px;
            }}
            .field textarea {{
                min-height: 110px;
                resize: vertical;
            }}
            .actions {{
                display: flex;
                gap: 8px;
                flex-wrap: wrap;
            }}
            .inline-form {{
                display: inline;
            }}
            .empty {{
                padding: 18px;
                background: white;
                border-radius: 14px;
                box-shadow: 0 8px 20px rgba(0,0,0,.06);
            }}
            .help {{
                color: #6b7280;
                font-size: 13px;
                margin-top: 8px;
            }}
            .alert {{
                background: #dcfce7;
                color: #166534;
                padding: 12px 14px;
                border-radius: 12px;
                margin-bottom: 16px;
            }}
            .alert-error {{
                background: #fee2e2;
                color: #991b1b;
            }}
        </style>
    </head>
    <body>
        <div class="layout">
            <aside class="sidebar">
                <div class="logo-box">{logo_html}</div>
                <div class="brand">
                    UPS System
                    <small>Ultra Power Solutions</small>
                </div>
                <div class="user-box">
                    <div><b>User:</b> {esc(user["username"])}</div>
                    <div><b>Role:</b> {esc(user["role"])}</div>
                </div>
                {nav_html(active, user["role"] == "admin")}
            </aside>
            <main class="content">
                <div class="page-title">{esc(title)}</div>
                {body}
            </main>
        </div>
    </body>
    </html>
    """


def get_row(section: str, row_id: int) -> sqlite3.Row:
    section_exists(section)
    table = SECTIONS[section]["table"]
    conn = get_conn()
    row = conn.execute(f"SELECT * FROM {table} WHERE id = ?", (row_id,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Record not found")
    return row


def dashboard_body() -> str:
    cards = []
    for key, cfg in SECTIONS.items():
        cards.append(
            f"""
            <div class="card">
                <div class="card-number">{count_rows(cfg["table"])}</div>
                <div>{esc(cfg["title"])}</div>
            </div>
            """
        )
    cards.append(
        f"""
        <div class="card">
            <div class="card-number">{count_rows("users")}</div>
            <div>Users</div>
        </div>
        """
    )
    return f'<div class="cards">{"".join(cards)}</div>'


def section_list_body(section: str) -> str:
    cfg = SECTIONS[section]
    table = cfg["table"]

    conn = get_conn()
    rows = conn.execute(f"SELECT * FROM {table} ORDER BY id DESC").fetchall()
    conn.close()

    headers = ["ID"] + [label for _, label, _ in cfg["fields"]] + ["Actions"]

    top = f"""
    <div class="toolbar">
        <a class="btn btn-primary" href="/ui/{section}/new">Add New</a>
        <a class="btn btn-success" href="/ui/{section}/import">Import Excel</a>
    </div>
    """

    if not rows:
        return top + '<div class="empty">No data found.</div>'

    body_rows = []
    for row in rows:
        cols = [f"<td>{row['id']}</td>"]
        for name, _, _ in cfg["fields"]:
            cols.append(f"<td>{esc(row[name])}</td>")
        cols.append(
            f"""
            <td>
                <div class="actions">
                    <a class="btn btn-warning" href="/ui/{section}/{row['id']}/edit">Edit</a>
                    <form class="inline-form" method="post" action="/ui/{section}/{row['id']}/delete">
                        <button class="btn btn-danger" type="submit">Delete</button>
                    </form>
                </div>
            </td>
            """
        )
        body_rows.append("<tr>" + "".join(cols) + "</tr>")

    head = "".join(f"<th>{esc(h)}</th>" for h in headers)
    return top + f"""
    <table>
        <thead><tr>{head}</tr></thead>
        <tbody>
            {"".join(body_rows)}
        </tbody>
    </table>
    """


def section_form_body(section: str, values: dict[str, Any] | None = None, edit_id: int | None = None) -> str:
    cfg = SECTIONS[section]
    values = values or {}
    title = f"Edit {cfg['singular']}" if edit_id else f"Add {cfg['singular']}"
    action = f"/ui/{section}/{edit_id}/edit" if edit_id else f"/ui/{section}/new"

    fields_html = []
    for name, label, field_type in cfg["fields"]:
        val = esc(values.get(name, ""))
        if field_type == "textarea":
            fields_html.append(
                f"""
                <div class="field">
                    <label>{esc(label)}</label>
                    <textarea name="{esc(name)}">{val}</textarea>
                </div>
                """
            )
        else:
            input_type = "number" if field_type == "number" else "text"
            step = ' step="0.01"' if field_type == "number" else ""
            fields_html.append(
                f"""
                <div class="field">
                    <label>{esc(label)}</label>
                    <input type="{input_type}" name="{esc(name)}" value="{val}"{step}>
                </div>
                """
            )

    return f"""
    <div class="form-box">
        <form method="post" action="{action}">
            {"".join(fields_html)}
            <div class="actions">
                <button class="btn btn-primary" type="submit">Save</button>
                <a class="btn btn-light" href="/ui/{section}">Back</a>
            </div>
        </form>
    </div>
    """


def import_page_body(section: str, message: str = "", error: str = "") -> str:
    cfg = SECTIONS[section]
    cols = ", ".join(name for name, _, _ in cfg["fields"])
    msg = f'<div class="alert">{esc(message)}</div>' if message else ""
    err = f'<div class="alert alert-error">{esc(error)}</div>' if error else ""

    return f"""
    {msg}
    {err}
    <div class="form-box">
        <form method="post" action="/ui/{section}/import" enctype="multipart/form-data">
            <div class="field">
                <label>Select Excel File</label>
                <input type="file" name="file" accept=".xlsx" required>
            </div>
            <div class="help">Excel columns must be exactly: {esc(cols)}</div>
            <div class="actions">
                <button class="btn btn-success" type="submit">Import</button>
                <a class="btn btn-light" href="/ui/{section}">Back</a>
            </div>
        </form>
    </div>
    """


def users_list_body() -> str:
    conn = get_conn()
    rows = conn.execute("SELECT * FROM users ORDER BY id DESC").fetchall()
    conn.close()

    top = """
    <div class="toolbar">
        <a class="btn btn-primary" href="/ui/users/new">Add New User</a>
    </div>
    """

    if not rows:
        return top + '<div class="empty">No users found.</div>'

    body_rows = []
    for row in rows:
        active_text = "Yes" if row["is_active"] == 1 else "No"
        body_rows.append(
            f"""
            <tr>
                <td>{row['id']}</td>
                <td>{esc(row['username'])}</td>
                <td>{esc(row['full_name'])}</td>
                <td>{esc(row['role'])}</td>
                <td>{active_text}</td>
                <td>
                    <div class="actions">
                        <a class="btn btn-warning" href="/ui/users/{row['id']}/edit">Edit</a>
                        <form class="inline-form" method="post" action="/ui/users/{row['id']}/delete">
                            <button class="btn btn-danger" type="submit">Delete</button>
                        </form>
                    </div>
                </td>
            </tr>
            """
        )

    return top + f"""
    <table>
        <thead>
            <tr>
                <th>ID</th>
                <th>Username</th>
                <th>Full Name</th>
                <th>Role</th>
                <th>Active</th>
                <th>Actions</th>
            </tr>
        </thead>
        <tbody>
            {"".join(body_rows)}
        </tbody>
    </table>
    """


def users_form_body(values: dict[str, Any] | None = None, edit_id: int | None = None) -> str:
    values = values or {}
    action = f"/ui/users/{edit_id}/edit" if edit_id else "/ui/users/new"
    checked = "checked" if str(values.get("is_active", 1)) == "1" else ""
    role = str(values.get("role", "user"))

    return f"""
    <div class="form-box">
        <form method="post" action="{action}">
            <div class="field">
                <label>Username</label>
                <input type="text" name="username" value="{esc(values.get('username', ''))}" required>
            </div>
            <div class="field">
                <label>Full Name</label>
                <input type="text" name="full_name" value="{esc(values.get('full_name', ''))}">
            </div>
            <div class="field">
                <label>Password</label>
                <input type="password" name="password" {'required' if not edit_id else ''}>
            </div>
            <div class="field">
                <label>Role</label>
                <select name="role">
                    <option value="admin" {'selected' if role == 'admin' else ''}>admin</option>
                    <option value="user" {'selected' if role == 'user' else ''}>user</option>
                </select>
            </div>
            <div class="field">
                <label><input type="checkbox" name="is_active" value="1" {checked}> Active</label>
            </div>
            <div class="actions">
                <button class="btn btn-primary" type="submit">Save</button>
                <a class="btn btn-light" href="/ui/users">Back</a>
            </div>
        </form>
    </div>
    """


def logo_page_body(message: str = "", error: str = "") -> str:
    msg = f'<div class="alert">{esc(message)}</div>' if message else ""
    err = f'<div class="alert alert-error">{esc(error)}</div>' if error else ""
    logo = logo_url()
    current = f'<img src="{logo}" style="max-width:260px;max-height:140px;object-fit:contain;">' if logo else "No logo uploaded."

    return f"""
    {msg}
    {err}
    <div class="form-box">
        <div class="field">
            <label>Current Logo</label>
            <div>{current}</div>
        </div>
        <form method="post" action="/ui/logo" enctype="multipart/form-data">
            <div class="field">
                <label>Select Logo Image</label>
                <input type="file" name="file" accept=".png,.jpg,.jpeg,.webp" required>
            </div>
            <div class="actions">
                <button class="btn btn-primary" type="submit">Upload Logo</button>
                <a class="btn btn-light" href="/ui">Back</a>
            </div>
        </form>
    </div>
    """


@app.get("/", response_class=HTMLResponse)
def root(request: Request):
    user = current_user(request)
    if user:
        return RedirectResponse(url="/ui", status_code=302)
    return RedirectResponse(url="/login", status_code=302)


@app.get("/login", response_class=HTMLResponse)
def login_get():
    return HTMLResponse(login_page_html())


@app.post("/login", response_class=HTMLResponse)
def login_post(username: str = Form(...), password: str = Form(...)):
    conn = get_conn()
    user = conn.execute(
        "SELECT * FROM users WHERE username = ? AND is_active = 1",
        (username.strip(),),
    ).fetchone()
    conn.close()

    if not user or user["password_hash"] != hash_password(password):
        return HTMLResponse(login_page_html("Invalid username or password"), status_code=401)

    response = RedirectResponse(url="/ui", status_code=303)
    response.set_cookie("ups_user", user["username"], httponly=True, samesite="lax")
    return response


@app.get("/logout")
def logout():
    response = RedirectResponse(url="/login", status_code=302)
    response.delete_cookie("ups_user")
    return response


@app.get("/ui", response_class=HTMLResponse)
def ui_dashboard(request: Request):
    user = require_user(request)
    return HTMLResponse(page_html("Dashboard", dashboard_body(), user, "dashboard"))


@app.get("/ui/logo", response_class=HTMLResponse)
def ui_logo(request: Request):
    user = require_admin(request)
    return HTMLResponse(page_html("Logo", logo_page_body(), user, "logo"))


@app.post("/ui/logo")
def ui_logo_post(request: Request, file: UploadFile = File(...)):
    user = require_admin(request)
    filename = file.filename or ""
    ext = os.path.splitext(filename)[1].lower()
    if ext not in [".png", ".jpg", ".jpeg", ".webp"]:
        return HTMLResponse(page_html("Logo", logo_page_body("", "Unsupported file type"), user, "logo"), status_code=400)

    save_path = os.path.join(UPLOAD_DIR, LOGO_FILENAME)
    with open(save_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    return HTMLResponse(page_html("Logo", logo_page_body("Logo uploaded successfully", ""), user, "logo"))


@app.get("/ui/users", response_class=HTMLResponse)
def ui_users(request: Request):
    user = require_admin(request)
    return HTMLResponse(page_html("Users", users_list_body(), user, "users"))


@app.get("/ui/users/new", response_class=HTMLResponse)
def ui_users_new(request: Request):
    user = require_admin(request)
    return HTMLResponse(page_html("Add User", users_form_body(), user, "users"))


@app.post("/ui/users/new")
def ui_users_new_post(
    request: Request,
    username: str = Form(...),
    full_name: str = Form(""),
    password: str = Form(...),
    role: str = Form("user"),
    is_active: str = Form("0"),
):
    require_admin(request)

    conn = get_conn()
    existing = conn.execute("SELECT id FROM users WHERE username = ?", (username.strip(),)).fetchone()
    if existing:
        conn.close()
        raise HTTPException(status_code=400, detail="Username already exists")

    conn.execute(
        "INSERT INTO users (username, full_name, password_hash, role, is_active) VALUES (?, ?, ?, ?, ?)",
        (
            username.strip(),
            full_name.strip(),
            hash_password(password),
            role.strip(),
            1 if is_active == "1" else 0,
        ),
    )
    conn.commit()
    conn.close()
    return RedirectResponse(url="/ui/users", status_code=303)


@app.get("/ui/users/{row_id}/edit", response_class=HTMLResponse)
def ui_users_edit(request: Request, row_id: int):
    user = require_admin(request)
    conn = get_conn()
    row = conn.execute("SELECT * FROM users WHERE id = ?", (row_id,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="User not found")
    return HTMLResponse(page_html("Edit User", users_form_body(dict(row), row_id), user, "users"))


@app.post("/ui/users/{row_id}/edit")
def ui_users_edit_post(
    request: Request,
    row_id: int,
    username: str = Form(...),
    full_name: str = Form(""),
    password: str = Form(""),
    role: str = Form("user"),
    is_active: str = Form("0"),
):
    require_admin(request)

    conn = get_conn()
    existing = conn.execute(
        "SELECT id FROM users WHERE username = ? AND id <> ?",
        (username.strip(), row_id),
    ).fetchone()
    if existing:
        conn.close()
        raise HTTPException(status_code=400, detail="Username already exists")

    if password.strip():
        conn.execute(
            "UPDATE users SET username = ?, full_name = ?, password_hash = ?, role = ?, is_active = ? WHERE id = ?",
            (
                username.strip(),
                full_name.strip(),
                hash_password(password),
                role.strip(),
                1 if is_active == "1" else 0,
                row_id,
            ),
        )
    else:
        conn.execute(
            "UPDATE users SET username = ?, full_name = ?, role = ?, is_active = ? WHERE id = ?",
            (
                username.strip(),
                full_name.strip(),
                role.strip(),
                1 if is_active == "1" else 0,
                row_id,
            ),
        )

    conn.commit()
    conn.close()
    return RedirectResponse(url="/ui/users", status_code=303)


@app.post("/ui/users/{row_id}/delete")
def ui_users_delete(request: Request, row_id: int):
    admin = require_admin(request)
    if admin["id"] == row_id:
        raise HTTPException(status_code=400, detail="You cannot delete yourself")

    conn = get_conn()
    conn.execute("DELETE FROM users WHERE id = ?", (row_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/ui/users", status_code=303)


@app.get("/ui/{section}", response_class=HTMLResponse)
def ui_section_list(request: Request, section: str):
    user = require_user(request)
    section_exists(section)
    return HTMLResponse(page_html(SECTIONS[section]["title"], section_list_body(section), user, section))


@app.get("/ui/{section}/new", response_class=HTMLResponse)
def ui_section_new(request: Request, section: str):
    user = require_user(request)
    section_exists(section)
    return HTMLResponse(page_html(f"Add {SECTIONS[section]['singular']}", section_form_body(section), user, section))


@app.post("/ui/{section}/new")
def ui_section_new_post(
    request: Request,
    section: str,
    name: str = Form(""),
    department: str = Form(""),
    phone: str = Form(""),
    address: str = Form(""),
    code: str = Form(""),
    account_type: str = Form(""),
    balance: float = Form(0),
    location: str = Form(""),
    status: str = Form(""),
    title: str = Form(""),
    client_name: str = Form(""),
    site_name: str = Form(""),
    description: str = Form(""),
    ticket_id: int = Form(0),
    from_location: str = Form(""),
    to_location: str = Form(""),
    cost: float = Form(0),
):
    require_user(request)
    section_exists(section)

    values_map = {
        "name": name,
        "department": department,
        "phone": phone,
        "address": address,
        "code": code,
        "account_type": account_type,
        "balance": balance,
        "location": location,
        "status": status,
        "title": title,
        "client_name": client_name,
        "site_name": site_name,
        "description": description,
        "ticket_id": ticket_id,
        "from_location": from_location,
        "to_location": to_location,
        "cost": cost,
    }

    cfg = SECTIONS[section]
    table = cfg["table"]
    columns = [field_name for field_name, _, _ in cfg["fields"]]
    values = [values_map[col] for col in columns]
    placeholders = ",".join("?" for _ in columns)

    conn = get_conn()
    conn.execute(
        f"INSERT INTO {table} ({','.join(columns)}) VALUES ({placeholders})",
        values,
    )
    conn.commit()
    conn.close()

    return RedirectResponse(url=f"/ui/{section}", status_code=303)


@app.get("/ui/{section}/{row_id}/edit", response_class=HTMLResponse)
def ui_section_edit(request: Request, section: str, row_id: int):
    user = require_user(request)
    row = get_row(section, row_id)
    return HTMLResponse(page_html(f"Edit {SECTIONS[section]['singular']}", section_form_body(section, dict(row), row_id), user, section))


@app.post("/ui/{section}/{row_id}/edit")
def ui_section_edit_post(
    request: Request,
    section: str,
    row_id: int,
    name: str = Form(""),
    department: str = Form(""),
    phone: str = Form(""),
    address: str = Form(""),
    code: str = Form(""),
    account_type: str = Form(""),
    balance: float = Form(0),
    location: str = Form(""),
    status: str = Form(""),
    title: str = Form(""),
    client_name: str = Form(""),
    site_name: str = Form(""),
    description: str = Form(""),
    ticket_id: int = Form(0),
    from_location: str = Form(""),
    to_location: str = Form(""),
    cost: float = Form(0),
):
    require_user(request)
    section_exists(section)

    values_map = {
        "name": name,
        "department": department,
        "phone": phone,
        "address": address,
        "code": code,
        "account_type": account_type,
        "balance": balance,
        "location": location,
        "status": status,
        "title": title,
        "client_name": client_name,
        "site_name": site_name,
        "description": description,
        "ticket_id": ticket_id,
        "from_location": from_location,
        "to_location": to_location,
        "cost": cost,
    }

    cfg = SECTIONS[section]
    table = cfg["table"]
    columns = [field_name for field_name, _, _ in cfg["fields"]]
    assignments = ", ".join(f"{col} = ?" for col in columns)
    values = [values_map[col] for col in columns] + [row_id]

    conn = get_conn()
    conn.execute(
        f"UPDATE {table} SET {assignments} WHERE id = ?",
        values,
    )
    conn.commit()
    conn.close()

    return RedirectResponse(url=f"/ui/{section}", status_code=303)


@app.post("/ui/{section}/{row_id}/delete")
def ui_section_delete(request: Request, section: str, row_id: int):
    require_user(request)
    section_exists(section)
    table = SECTIONS[section]["table"]

    conn = get_conn()
    conn.execute(f"DELETE FROM {table} WHERE id = ?", (row_id,))
    conn.commit()
    conn.close()

    return RedirectResponse(url=f"/ui/{section}", status_code=303)


@app.get("/ui/{section}/import", response_class=HTMLResponse)
def ui_section_import(request: Request, section: str):
    user = require_user(request)
    section_exists(section)
    return HTMLResponse(page_html(f"Import {SECTIONS[section]['title']}", import_page_body(section), user, section))


@app.post("/ui/{section}/import", response_class=HTMLResponse)
def ui_section_import_post(request: Request, section: str, file: UploadFile = File(...)):
    user = require_user(request)
    section_exists(section)
    cfg = SECTIONS[section]
    table = cfg["table"]
    expected_cols = [field_name for field_name, _, _ in cfg["fields"]]

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        return HTMLResponse(page_html(f"Import {cfg['title']}", import_page_body(section, "", "Please upload .xlsx file"), user, section), status_code=400)

    temp_path = os.path.join(UPLOAD_DIR, f"temp_{section}.xlsx")
    with open(temp_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    try:
        wb = openpyxl.load_workbook(temp_path)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]

        if headers != expected_cols:
            os.remove(temp_path)
            return HTMLResponse(
                page_html(
                    f"Import {cfg['title']}",
                    import_page_body(section, "", f"Excel headers must be exactly: {', '.join(expected_cols)}"),
                    user,
                    section,
                ),
                status_code=400,
            )

        conn = get_conn()
        imported = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            values = list(row[:len(expected_cols)])
            values = ["" if v is None else v for v in values]
            placeholders = ",".join("?" for _ in expected_cols)
            conn.execute(
                f"INSERT INTO {table} ({','.join(expected_cols)}) VALUES ({placeholders})",
                values,
            )
            imported += 1

        conn.commit()
        conn.close()
        os.remove(temp_path)

        return HTMLResponse(
            page_html(
                f"Import {cfg['title']}",
                import_page_body(section, f"Imported {imported} rows successfully", ""),
                user,
                section,
            )
        )
    except Exception as e:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        return HTMLResponse(
            page_html(
                f"Import {cfg['title']}",
                import_page_body(section, "", f"Import failed: {str(e)}"),
                user,
                section,
            ),
            status_code=500,
        )